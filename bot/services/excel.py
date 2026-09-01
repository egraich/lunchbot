"""Месячный xlsx-отчёт в формате бумажной таблицы классрука."""

import os
import tempfile
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from bot.db import records, students
from bot.db.admins import Admin

CELL_LABEL = {"O": "О", "O1": "О(1)"}
MONTH_NAMES = (
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
)

# символы, запрещённые в названии листа Excel и опасные в имени файла
_BAD_CHARS = "[]:*?/\\"


def _safe(value: str) -> str:
    for ch in _BAD_CHARS:
        value = value.replace(ch, "-")
    return value.strip() or "class"


def month_weekdays(year: int, month: int) -> list[date]:
    """Все пн–пт месяца — колонки отчёта (выходных колонок нет, как на фото)."""
    d = date(year, month, 1)
    days: list[date] = []
    while d.month == month:
        if d.weekday() < 5:
            days.append(d)
        d += timedelta(days=1)
    return days


def generate(
    path: str,
    class_name: str,
    year: int,
    month: int,
    students_list: list[tuple[int, str]],
    records_map: dict[tuple[int, str], str],
) -> None:
    days = month_weekdays(year, month)
    wb = Workbook()
    ws = wb.active
    ws.title = _safe(class_name)[:31]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    bold_italic = Font(bold=True, italic=True)

    ws.cell(row=1, column=1, value=f"{MONTH_NAMES[month - 1]}:").font = bold_italic
    for col, d in enumerate(days, start=2):
        ws.cell(row=1, column=col, value=d.day).alignment = center

    row = 2
    for sid, name in students_list:
        ws.cell(row=row, column=1, value=name)
        for col, d in enumerate(days, start=2):
            status = records_map.get((sid, d.isoformat()))
            if status:
                ws.cell(row=row, column=col, value=CELL_LABEL[status]).alignment = center
        row += 1

    for label, wanted in (("Без первого :", "O"), ("С первым :", "O1")):
        ws.cell(row=row, column=1, value=label).font = bold_italic
        for col, d in enumerate(days, start=2):
            n = sum(1 for sid, _ in students_list if records_map.get((sid, d.isoformat())) == wanted)
            ws.cell(row=row, column=col, value=n).alignment = center
        row += 1

    last_col = get_column_letter(len(days) + 1)
    for cells in ws[f"A1:{last_col}{row - 1}"]:
        for cell in cells:
            cell.border = border

    ws.column_dimensions["A"].width = 16
    for col in range(2, len(days) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    wb.save(path)


async def build_report(admin: Admin, year: int, month: int) -> Path | None:
    """Собрать отчёт по классу админа во временный файл. None — если нет учеников."""
    # удалённые посреди месяца ученики остаются в отчёте, если успели поесть
    start = f"{year:04d}-{month:02d}-01"
    end = f"{year:04d}-{month:02d}-31"
    sts = await students.list_for_report(admin.school_id, admin.class_name, start, end)
    if not sts:
        return None
    recs = await records.for_month(admin.school_id, admin.class_name, year, month)
    fd, raw = tempfile.mkstemp(suffix=".xlsx", prefix=f"{_safe(admin.class_name)}_{year}-{month:02d}_")
    os.close(fd)
    generate(raw, admin.class_name, year, month, [(s.id, s.name) for s in sts], recs)
    return Path(raw)
