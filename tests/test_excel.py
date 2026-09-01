"""Unit tests for the Excel report: format O/O(1)/empty + sums."""

from openpyxl import load_workbook

from bot.services.excel import generate, month_weekdays


def test_month_weekdays_excludes_weekends():
    days = month_weekdays(2025, 12)
    assert [d.day for d in days[:5]] == [1, 2, 3, 4, 5]
    assert all(d.weekday() < 5 for d in days)
    assert 6 not in [d.day for d in days]
    assert 7 not in [d.day for d in days]
    assert 8 in [d.day for d in days]


def test_report_layout(tmp_path):
    students = [(1, "Старик Ф"), (2, "Климович И")]
    records = {
        (1, "2025-12-01"): "O",
        (2, "2025-12-01"): "O1",
        (1, "2025-12-02"): "O",
    }
    path = tmp_path / "report.xlsx"
    generate(str(path), "11-Б", 2025, 12, students, records)
    ws = load_workbook(path).active

    assert ws.cell(1, 1).value == "Декабрь:"
    header = [ws.cell(1, c).value for c in range(2, 2 + len(month_weekdays(2025, 12)))]
    assert header[0] == 1 and 6 not in header and 7 not in header and 8 in header

    assert ws.cell(2, 1).value == "Старик Ф"
    assert ws.cell(2, 2).value == "О"
    assert ws.cell(3, 2).value == "О(1)"
    assert ws.cell(2, 4).value is None

    summary = 2 + len(students)
    assert ws.cell(summary, 1).value == "Без первого :"
    assert ws.cell(summary, 2).value == 1
    assert ws.cell(summary + 1, 1).value == "С первым :"
    assert ws.cell(summary + 1, 2).value == 1
    assert ws.cell(summary, 4).value == 0
